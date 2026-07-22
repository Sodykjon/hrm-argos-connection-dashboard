# -*- coding: utf-8 -*-
"""Generate seed-snapshot.json + registry.json from real ARGOS xlsx files.
Mirrors the TS runtime parser shape exactly."""
import sys, io, json, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

DL = "Downloads/Telegram Desktop"
HISOBOT = os.path.join(DL, "HRM_ARGOS_hisobot_02072026 (1) (2).xlsx")
REESTR  = os.path.join(DL, "1700 РЕЕСТР_02.07.2026.xlsx")
OUT = os.path.dirname(os.path.abspath(__file__))

def norm_stir(v):
    if v is None: return ""
    s = str(v).strip()
    if s.endswith(".0"): s = s[:-2]
    return s

def norm_status(v):
    s = (str(v) if v is not None else "").strip().lower()
    if s == "faol": return "ulangan"
    if "chiril" in s or "ochiril" in s: return "ochirilgan"   # tizimdan o`chirildi
    if "ланмаган" in s or s == "уланмаган": return "ulanmagan"
    return "ulanmagan"  # safe default (only Уланмаган lands here)

# ---------- snapshot ----------
wb = openpyxl.load_workbook(HISOBOT, read_only=True, data_only=True)
ws = wb["Ҳисобот"]
srows = list(ws.iter_rows(values_only=True))
# region order + summary numbers from Ҳисобот sheet
summary = {}   # region -> dict from sheet
order = []
title = ""
for r in srows:
    vals = [ (str(c).strip() if c is not None else "") for c in r ]
    for v in vals:
        if "ҳолатига" in v: title = v
    # detect data rows: has region name in col idx 2 and numeric cols 3..7
    name = vals[2] if len(vals) > 2 else ""
    if name and name != "ЖАМИ" and name != "ҲУДУД НОМИ":
        try:
            total = int(float(vals[3])); ul = int(float(vals[4]))
            un = int(float(vals[5])); och = int(float(vals[6]))
        except (ValueError, IndexError):
            continue
        summary[name] = dict(total=total, ulangan=ul, ulanmagan=un, ochirilgan=och)
        order.append(name)

m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", title)
date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else "2026-07-02"

ws = wb["Маълумотлар"]
drows = list(ws.iter_rows(values_only=True))
orgs = []
for r in drows[1:]:
    region = (str(r[0]).strip() if r[0] is not None else "")
    name   = (str(r[1]).strip() if r[1] is not None else "")
    stir   = norm_stir(r[2])
    status = norm_status(r[3])
    contract = (str(r[4]).strip() if r[4] is not None else "")
    if not name and not stir: continue
    orgs.append(dict(region=region, name=name, stir=stir, status=status, contract=contract))
wb.close()

# aggregate regions from orgs
from collections import defaultdict
agg = defaultdict(lambda: dict(total=0, ulangan=0, ulanmagan=0, ochirilgan=0))
for o in orgs:
    a = agg[o["region"]]
    a["total"] += 1
    a[o["status"]] += 1

# build regions in sheet order
regions = []
for name in order:
    a = agg.get(name, dict(total=0, ulangan=0, ulanmagan=0, ochirilgan=0))
    pct = (a["ulangan"]/a["total"]) if a["total"] else 0
    regions.append(dict(name=name, total=a["total"], ulangan=a["ulangan"],
                        ulanmagan=a["ulanmagan"], ochirilgan=a["ochirilgan"], percent=round(pct,6)))

t_total = sum(r["total"] for r in regions)
t_ul = sum(r["ulangan"] for r in regions)
t_un = sum(r["ulanmagan"] for r in regions)
t_och = sum(r["ochirilgan"] for r in regions)
totals = dict(total=t_total, ulangan=t_ul, ulanmagan=t_un, ochirilgan=t_och,
              percent=round(t_ul/t_total,6) if t_total else 0)

snapshot = dict(date=date, uploadedAt=date+"T00:00:00.000Z", totals=totals, regions=regions, orgs=orgs)
with open(os.path.join(OUT, "seed-snapshot.json"), "w", encoding="utf-8") as f:
    json.dump(snapshot, f, ensure_ascii=False, separators=(",",":"))

# ---------- validation vs summary sheet ----------
print("=== VALIDATION (agg from orgs vs Ҳисобот sheet) ===")
print(f"date={date}  orgs={len(orgs)}")
print(f"TOTALS agg: {totals}")
ok = True
for name in order:
    a = agg.get(name)
    s = summary[name]
    match = a and a["total"]==s["total"] and a["ulangan"]==s["ulangan"] and a["ulanmagan"]==s["ulanmagan"] and a["ochirilgan"]==s["ochirilgan"]
    if not match:
        ok = False
        print(f"  MISMATCH {name}: agg={dict(a) if a else None} sheet={s}")
print("ALL REGIONS MATCH SHEET ✓" if ok else "  ^^ mismatches above")
grand = dict(total=3886, ulangan=2599, ulanmagan=1012, ochirilgan=275)
print("GRAND vs known 3886/2599/1012/275:",
      totals["total"]==grand["total"], totals["ulangan"]==grand["ulangan"],
      totals["ulanmagan"]==grand["ulanmagan"], totals["ochirilgan"]==grand["ochirilgan"])

# ---------- registry ----------
wb = openpyxl.load_workbook(REESTR, read_only=True, data_only=True)
reg = {}
ws = wb["ум.реестр (2)"]
for r in ws.iter_rows(min_row=5, values_only=True):
    stir = norm_stir(r[10]) if len(r) > 10 else ""
    if not stir: continue
    def g(i): return (str(r[i]).strip() if len(r) > i and r[i] is not None else "")
    tel = g(12) or g(13)
    entry = dict(name=g(1), rahbar=g(7), manzil=g(4), email=g(11), tel=tel, mhobt=g(20))
    if stir not in reg:
        reg[stir] = entry
# merge tulov% from directory ИНН
ws = wb["directory ИНН"]
for r in ws.iter_rows(min_row=2, values_only=True):
    stir = norm_stir(r[1]) if len(r) > 1 else ""
    if not stir: continue
    tulov = (str(r[4]).strip() if len(r) > 4 and r[4] is not None else "")
    if stir in reg:
        reg[stir]["tulov"] = tulov
    else:
        reg[stir] = dict(name="", rahbar="", manzil="", email="", tel="", mhobt="", tulov=tulov)
wb.close()
with open(os.path.join(OUT, "registry.json"), "w", encoding="utf-8") as f:
    json.dump(reg, f, ensure_ascii=False, separators=(",",":"))

# coverage: how many ulanmagan orgs have a registry match?
un_orgs = [o for o in orgs if o["status"]=="ulanmagan"]
matched = sum(1 for o in un_orgs if o["stir"] in reg)
print(f"\n=== REGISTRY ===  entries={len(reg)}")
print(f"ulanmagan orgs={len(un_orgs)}, matched in registry={matched} ({matched*100//max(1,len(un_orgs))}%)")
allmatched = sum(1 for o in orgs if o['stir'] in reg)
print(f"ALL orgs matched in registry={allmatched}/{len(orgs)}")
print("\nWROTE:", os.path.join(OUT,"seed-snapshot.json"), "and registry.json")
print("seed size KB:", os.path.getsize(os.path.join(OUT,"seed-snapshot.json"))//1024,
      "| registry KB:", os.path.getsize(os.path.join(OUT,"registry.json"))//1024)
